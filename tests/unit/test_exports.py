import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from bot_atul.db.migrations import migrate
from bot_atul.db.repositories import Repository
from bot_atul.services.exports import export_tickets


def test_export_has_four_sheets_and_lossless_long_description(tmp_path: Path) -> None:
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    migrate(connection)
    repository = Repository(connection)
    repository.upsert_user(10, "reporter")
    description = "🚀abc" * 10_000
    ticket = repository.create_ticket(
        reporter_id=10,
        service_name="Other",
        urgency="Critical",
        title="Render failure",
        description=description,
    )
    repository.save_dashboard_card(ticket.number, 101)

    path = export_tickets(repository, tmp_path / "issues.xlsx", -1001, 24)
    workbook = load_workbook(path)

    assert workbook.sheetnames == [
        "Issues",
        "Summary",
        "Status History",
        "Description Parts",
    ]
    issues = workbook["Issues"]
    assert issues.auto_filter.ref is not None
    assert issues.freeze_panes == "A2"
    assert issues["F2"].value == "Render failure"
    assert "Description Parts" in issues["G2"].value
    parts = workbook["Description Parts"]
    reconstructed = "".join(
        row[2] for row in parts.iter_rows(min_row=2, values_only=True)
    )
    assert reconstructed == description
    assert issues["P2"].hyperlink is not None
    assert issues["P2"].hyperlink.target == "https://t.me/c/1/24/101"
