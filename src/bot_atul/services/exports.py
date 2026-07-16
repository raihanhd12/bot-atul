from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from sqlite3 import Row, connect
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from bot_atul.db.repositories import Repository
from bot_atul.services.dashboard import topic_link

CELL_LIMIT = 32_767
PART_SIZE = 32_000


def export_tickets(
    repository: Repository,
    path: Path,
    team_group_id: int,
    dashboard_topic_id: int,
    start: date | None = None,
    end: date | None = None,
) -> Path:
    snapshot = connect(":memory:")
    snapshot.row_factory = repository.connection.row_factory
    repository.connection.backup(snapshot)
    repository = Repository(snapshot)
    where: list[str] = []
    params: list[str] = []
    if start:
        where.append("date(t.created_at, '+7 hours') >= ?")
        params.append(start.isoformat())
    if end:
        where.append("date(t.created_at, '+7 hours') <= ?")
        params.append(end.isoformat())
    clause = f"WHERE {' AND '.join(where)}" if where else ""
    tickets = repository.connection.execute(
        f"""
        SELECT t.number, datetime(t.created_at, '+7 hours') AS created_at,
               datetime(t.updated_at, '+7 hours') AS updated_at,
               t.status, t.urgency, t.service_name,
               t.title, t.description, t.reporter_id, t.assignee_id,
               t.fixed_at, t.closed_at,
               t.topic_id, c.message_id AS dashboard_card_id
        FROM tickets t
        LEFT JOIN ticket_dashboard_cards c ON c.ticket_number = t.number
        {clause} ORDER BY t.number
        """,
        params,
    ).fetchall()

    workbook = Workbook()
    issues = cast(Worksheet, workbook.active)
    issues.title = "Issues"
    summary = workbook.create_sheet("Summary")
    history = workbook.create_sheet("Status History")
    parts = workbook.create_sheet("Description Parts")
    issue_headers = [
        "Ticket",
        "Created",
        "Updated",
        "Status",
        "Urgency",
        "Title",
        "Description",
        "Service",
        "Reporter",
        "Assignee",
        "Fixed",
        "Closed",
        "Age (days)",
        "Dashboard Card ID",
        "Description Parts",
        "Ticket Link",
    ]
    issues.append(issue_headers)
    parts.append(["Ticket", "Part", "Text"])

    for row in tickets:
        description = str(row["description"])
        split = [
            description[index : index + PART_SIZE]
            for index in range(0, len(description), PART_SIZE)
        ]
        long_description = len(description) > CELL_LIMIT
        if long_description:
            for index, chunk in enumerate(split, start=1):
                parts.append([row["number"], index, chunk])
        link = (
            topic_link(
                team_group_id,
                dashboard_topic_id,
                int(row["dashboard_card_id"]),
            )
            if row["dashboard_card_id"] is not None
            else (
                topic_link(team_group_id, int(row["topic_id"]))
                if row["topic_id"] is not None
                else None
            )
        )
        issues.append(
            [
                row["number"],
                datetime.fromisoformat(row["created_at"]),
                datetime.fromisoformat(row["updated_at"]),
                row["status"],
                row["urgency"],
                row["title"],
                (
                    f"See Description Parts rows for ticket #{row['number']}"
                    if long_description
                    else description
                ),
                row["service_name"],
                str(row["reporter_id"]),
                str(row["assignee_id"]) if row["assignee_id"] else None,
                _as_datetime(row["fixed_at"]),
                _as_datetime(row["closed_at"]),
                None,
                (
                    str(row["dashboard_card_id"])
                    if row["dashboard_card_id"]
                    else None
                ),
                len(split) if long_description else 0,
                link,
            ]
        )
        if link:
            cell = issues.cell(issues.max_row, 16)
            cell.hyperlink = link
            cell.style = "Hyperlink"

    _write_summary(summary, tickets)
    _write_history(repository, history, [int(row["number"]) for row in tickets])
    for sheet in workbook.worksheets:
        _style_sheet(sheet)
    for row in issues.iter_rows(min_row=2):
        for index in (2, 3, 11, 12):
            row[index - 1].number_format = "yyyy-mm-dd hh:mm"
        row[6].alignment = Alignment(wrap_text=False, vertical="top")
        row_number = row[0].row
        assert row_number is not None
        issues.row_dimensions[row_number].height = 24
        status_colors = {
            "Open": "FDECEC",
            "In Progress": "FFF4D6",
            "Fixed": "E8F7EE",
            "Closed": "E7EEF7",
        }
        row[3].fill = PatternFill(
            "solid", fgColor=status_colors.get(str(row[3].value), "FFFFFF")
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    snapshot.close()
    return path


def _write_summary(sheet: Worksheet, tickets: list[Row]) -> None:
    sheet.append(["Group", "Value", "Count"])
    for group, key in (("Status", "status"), ("Service", "service_name")):
        counts = Counter(str(row[key]) for row in tickets)
        for value, count in sorted(counts.items()):
            sheet.append([group, value, count])
    counts = Counter(str(row["created_at"])[:10] for row in tickets)
    for value, count in sorted(counts.items()):
        sheet.append(["Created Date", value, count])


def _write_history(
    repository: Repository, sheet: Worksheet, ticket_numbers: list[int]
) -> None:
    sheet.append(["Ticket", "Previous", "New", "Actor", "Reason", "Timestamp"])
    if not ticket_numbers:
        return
    placeholders = ",".join("?" for _ in ticket_numbers)
    rows = repository.connection.execute(
        f"""
        SELECT ticket_number, previous_status, new_status, actor_id, reason, created_at
        FROM status_history WHERE ticket_number IN ({placeholders}) ORDER BY id
        """,
        ticket_numbers,
    ).fetchall()
    for row in rows:
        sheet.append(list(row))


def _style_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_view.zoomScale = 85
    sheet.print_title_rows = "1:1"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="16324F")
    for column in sheet.columns:
        column_number = column[0].column
        assert column_number is not None
        letter = get_column_letter(column_number)
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 50)
        sheet.column_dimensions[letter].width = width


def _as_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    return datetime.fromisoformat(value) + timedelta(hours=7)
